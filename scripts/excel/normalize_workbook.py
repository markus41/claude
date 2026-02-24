#!/usr/bin/env python3
"""Normalize workbook field names and string content for downstream export."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def normalize_header(value: object) -> object:
    if value is None:
        return value
    return str(value).strip().lower().replace(" ", "_")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    wb = load_workbook(args.input)
    try:
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.value = normalize_header(cell.value)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = cell.value.strip()
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        wb.save(args.output)
    finally:
        wb.close()


if __name__ == "__main__":
    main()
