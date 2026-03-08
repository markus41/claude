"""Workbook validation engine for TAIA quality gates."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from .parsers import parse_sheet_in_chunks, profile_snapshot
from .schema_profiles import WorkbookProfile, get_profile


CHECK_WEIGHTS = {
    "required_sheets": 0.25,
    "named_ranges": 0.2,
    "formula_integrity": 0.2,
    "hidden_sheet_policy": 0.1,
    "merged_cell_policy": 0.1,
    "type_drift": 0.15,
}


def _value_type(value) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    return "string"


def _validate_types(workbook_path: Path, profile: WorkbookProfile) -> List[str]:
    failures: List[str] = []
    for sheet_name, rules in profile.columns_by_sheet.items():
        expected_by_col = {rule.name: rule.expected_type for rule in rules}
        type_counts = defaultdict(lambda: defaultdict(int))

        for chunk in parse_sheet_in_chunks(workbook_path, sheet_name, chunk_size=250):
            for row in chunk.rows:
                for col, expected in expected_by_col.items():
                    value = row.get(col)
                    observed = _value_type(value)
                    if observed != "empty":
                        type_counts[col][observed] += 1
                    if expected != observed and observed != "empty":
                        failures.append(
                            f"type drift in {sheet_name}.{col} row~{chunk.row_offset}: expected {expected}, observed {observed}"
                        )

        for col, observed_counts in type_counts.items():
            if not observed_counts:
                continue
            dominant = max(observed_counts, key=observed_counts.get)
            if dominant != expected_by_col[col]:
                failures.append(
                    f"dominant type drift in {sheet_name}.{col}: expected {expected_by_col[col]}, observed {dominant}"
                )
    return failures


def validate_workbook(workbook_path: str | Path, profile_id: str) -> Dict[str, object]:
    path = Path(workbook_path)
    profile = get_profile(profile_id)
    snapshot = profile_snapshot(path)

    workbook = load_workbook(filename=str(path), read_only=True, data_only=False)
    try:
        failures: Dict[str, List[str]] = defaultdict(list)
        score_breakdown: Dict[str, float] = {}

        missing_sheets = sorted(set(profile.required_sheets) - set(snapshot["sheet_names"]))
        if missing_sheets:
            failures["required_sheets"].append(f"missing sheets: {', '.join(missing_sheets)}")
        score_breakdown["required_sheets"] = 1.0 if not missing_sheets else 0.0

        missing_names = sorted(set(profile.required_named_ranges) - set(snapshot["defined_names"]))
        if missing_names:
            failures["named_ranges"].append(f"missing named ranges: {', '.join(missing_names)}")
        score_breakdown["named_ranges"] = 1.0 if not missing_names else 0.0

        formula_failures: List[str] = []
        for sheet_name, rules in profile.columns_by_sheet.items():
            if sheet_name not in workbook.sheetnames:
                continue
            expected_formula_cols = {r.name for r in rules if r.expected_type == "formula"}
            ws = workbook[sheet_name]
            header_cells = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            header_index = {str(name): idx for idx, name in enumerate(header_cells) if name is not None}
            for formula_col in expected_formula_cols:
                idx = header_index.get(formula_col)
                if idx is None:
                    formula_failures.append(f"{sheet_name}.{formula_col} missing from header")
                    continue
                sample = next(ws.iter_rows(min_row=2, min_col=idx + 1, max_col=idx + 1, max_row=2), None)
                if sample and (sample[0].value is None or not str(sample[0].value).startswith("=")):
                    formula_failures.append(f"{sheet_name}.{formula_col} row 2 missing formula")
        if formula_failures:
            failures["formula_integrity"].extend(formula_failures)
        score_breakdown["formula_integrity"] = 1.0 if not formula_failures else 0.0

        hidden = snapshot["hidden_sheets"]
        if hidden:
            failures["hidden_sheet_policy"].append(f"hidden sheets present: {', '.join(hidden)}")
        score_breakdown["hidden_sheet_policy"] = 1.0 if not hidden else 0.0

        merged_failures = [
            f"{sheet}: merged cells {', '.join(ranges)}"
            for sheet, ranges in snapshot["merged_cell_ranges"].items()
            if ranges
        ]
        if merged_failures:
            failures["merged_cell_policy"].extend(merged_failures)
        score_breakdown["merged_cell_policy"] = 1.0 if not merged_failures else 0.0

        type_drift_failures = _validate_types(path, profile)
        if type_drift_failures:
            failures["type_drift"].extend(type_drift_failures[:50])
        score_breakdown["type_drift"] = 1.0 if not type_drift_failures else 0.0

        weighted_score = sum(score_breakdown[name] * CHECK_WEIGHTS[name] for name in CHECK_WEIGHTS)
        return {
            "profile": asdict(profile),
            "checks": score_breakdown,
            "weighted_score": round(weighted_score * 100, 2),
            "failures": dict(failures),
            "passed": not failures,
            "snapshot": snapshot,
        }
    finally:
        workbook.close()
