"""Chunked, memory-safe workbook parsing utilities."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Generator, Iterable, List

from openpyxl import load_workbook


@dataclass
class SheetChunk:
    sheet_name: str
    rows: List[Dict[str, object]]
    row_offset: int


def _normalize_headers(headers: Iterable[object]) -> List[str]:
    normalized: List[str] = []
    for idx, header in enumerate(headers):
        label = str(header).strip() if header is not None else ""
        normalized.append(label or f"column_{idx + 1}")
    return normalized


def parse_sheet_in_chunks(workbook_path: str | Path, sheet_name: str, chunk_size: int = 500) -> Generator[SheetChunk, None, None]:
    """Yield row dictionaries in chunk-sized batches to avoid loading entire sheets."""
    workbook = load_workbook(filename=str(workbook_path), read_only=True, data_only=False)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = _normalize_headers(next(rows, []))
        batch: List[Dict[str, object]] = []
        start_row = 2
        row_cursor = 2

        for row in rows:
            record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            batch.append(record)
            if len(batch) >= chunk_size:
                yield SheetChunk(sheet_name=sheet_name, rows=batch, row_offset=start_row)
                batch = []
                start_row = row_cursor + 1
            row_cursor += 1

        if batch:
            yield SheetChunk(sheet_name=sheet_name, rows=batch, row_offset=start_row)
    finally:
        workbook.close()


def profile_snapshot(workbook_path: str | Path) -> Dict[str, object]:
    """Capture a workbook metadata snapshot without materializing all rows."""
    workbook = load_workbook(filename=str(workbook_path), read_only=True, data_only=False)
    try:
        return {
            "sheet_names": list(workbook.sheetnames),
            "defined_names": sorted(name for name in workbook.defined_names.keys()),
            "hidden_sheets": [
                ws.title
                for ws in workbook.worksheets
                if getattr(ws, "sheet_state", "visible") != "visible"
            ],
            "merged_cell_ranges": {
                ws.title: [str(rng) for rng in ws.merged_cells.ranges]
                for ws in workbook.worksheets
            },
        }
    finally:
        workbook.close()
